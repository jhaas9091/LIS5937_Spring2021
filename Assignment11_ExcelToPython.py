"""Module 10 Assignment
Julie Haas
LIS 5937
Spring 2021"""

from openpyxl import load_workbook
import pandas as pd

from Assignment11_classes import Covid, Ages
from Assignment11_mapping import COUNTY, COUNTYNAME, PUIsTotal, Age_0_9, Age_10_19, Age_20_29, \
    Age_30_39, Age_40_49, Age_50_59, Age_60_69, Age_70_79, Age_80plus, Age_Unkn

# Read and sort data from original Excel sheet, then write it to new file
df = pd.read_excel(r'COVID19_03242020_ByCounty.xlsx', usecols ='E:Y')
workbook = pd.DataFrame(df)
sorted_workbook = workbook.sort_values(by='COUNTYNAME')
sorted_workbook.to_excel("COVID19_03242020_ByCounty_Sorted.xlsx", sheet_name = 'SortedByCounty')

# Read the sorted worksheet
workbook = load_workbook(filename='COVID19_03242020_ByCounty_Sorted.xlsx', read_only=False)
sheet = workbook.active

pui_s = []
age_s = []

# Using values_only because we want to return the cell value
for row in sheet.iter_rows(min_row=2, values_only=True):
    pui = Covid(County=row[COUNTY], CountyName=row[COUNTYNAME], PUIsTotal=row[PUIsTotal])
    pui_s.append(pui)

    age = Ages(ages0to9= row[Age_0_9],
                    ages10to19 = row[Age_10_19],
                    ages20to29 = row[Age_20_29],
                    ages30to39 = row[Age_30_39],
                    ages40to49 = row[Age_40_49],
                    ages50to59 = row[Age_50_59],
                    ages60to69 = row[Age_60_69],
                    ages70to79 = row[Age_70_79],
                    ages80plus = row[Age_80plus],
                    agesUnknown = row[Age_Unkn])
    age_s.append(age)

# Iterate through each county and print out data
for i in range(0, 68):
    if i < 68:
        print(pui_s[i])
        #print(age_s[i])
    else:
        break

